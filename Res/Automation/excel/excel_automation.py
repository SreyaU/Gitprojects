import openpyxl
# import statistics

'''create & write to excel file'''
wb = openpyxl.Workbook()

sheet = wb["Sheet"]
wb.remove(sheet)

wb.create_sheet("dravid")
wb.create_sheet("sachin")

sheet = wb["dravid"]
sheet.cell(row = 1, column = 3).value = 7777
sheet = wb["sachin"]
# sheet.cell(row = 3, column = 7).value = 9999




'''update data'''
import os
filenames = []
for filename in os.listdir():
    if filename.endswith(".txt"):
        filenames.append(filename)

for filename in filenames:
    player = filename.replace(".txt","")

    sheet = wb.create_sheet(player)
    fobj =open(filename,'r')
    for line_number, line in enumerate(fobj):
        columns = line.split()

        runs = int(columns[2])
        balls = int(columns[3])

        sheet.cell(row = line_number + 1, column =1).value = runs
        sheet.cell(row = line_number + 1, column =2).value = balls
wb.save("cricket.xlsx")


# '''read from excel file'''
# wb = openpyxl.load_workbook("cricket.xlsx")









